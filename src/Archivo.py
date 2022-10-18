import string
import os
from os import remove
from os import path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import stylesheet
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from openpyxl.utils import get_column_letter
from datetime import datetime

import openpyxl.styles

import time





class Archivo:

    def __init__(self):
        self.data_completa = []

    def archivandoDatos(self, tipo_de_cambio, tasa_diaria, bolsa_de_valores):

        if path.exists("Indicadores_financieros.xlsx"):
            remove('Indicadores_financieros.xlsx')

        #fecha = datetime.today().strftime('%Y-%m-%d')
        #self.data_completa = [fecha] + tipo_de_cambio + tasa_diaria + bolsa_de_valores
        self.data_completa = tipo_de_cambio + tasa_diaria + bolsa_de_valores

        wb = Workbook()
        sheet = wb.active



        sheet['B3'] = 'COMPRA'
        sheet['C3'] = 'VENTA'
        sheet['D3'] = 'COMPRA'
        sheet['E3'] = 'VENTA'
        sheet.merge_cells('F2:F3')
        sheet['F2'] = 'TAMN'
        sheet.merge_cells('G2:G3')
        sheet['G2'] = 'TAMEX'
        sheet.merge_cells('H2:H3')
        sheet['H2'] = 'Índice General Bursátil Diaria (Var. %)'
        sheet.merge_cells('I2:I3')
        sheet['I2'] = 'Índice Selectivo Bursátil Diaria (Var. %)'
        sheet.merge_cells('J2:J3')
        sheet['J2'] = 'Monto negociado en acciones (Mill. S/.) - Prom. Diario'
        sheet.merge_cells('K2:K3')
        sheet['K2'] = 'Índice General Bursátil Mensual (Var. %)'
        sheet.merge_cells('L2:L3')
        sheet['L2'] = 'Índice Selectivo Bursátil Mensual (Var. %)'
        sheet.merge_cells('M2:M3')
        sheet['M2'] = 'Indicadores bursátiles - IGB (índice)'
        sheet.merge_cells('N2:N3')
        sheet['N2'] = 'Indicadores bursátiles - ISB (índice)'

        sheet.merge_cells('B1:G1')
        sheet['B1'] = 'SBS'
        sheet['B1'].fill = PatternFill(start_color="FFED00", fill_type="solid")
        sheet['B1'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('H1:N1')
        sheet['H1'] = 'BVL'
        sheet['H1'].fill = PatternFill(start_color="0082FF", fill_type="solid")
        sheet['H1'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('B2:C2')
        sheet['B2'] = 'DOLAR'
        sheet['B2'].fill = PatternFill(start_color="84bd00", fill_type="solid")
        sheet['B2'].alignment = Alignment(horizontal='center')

        sheet.merge_cells('D2:E2')
        sheet['D2'] = 'EURO'
        sheet['D2'].fill = PatternFill(start_color="f08e06", fill_type="solid")
        sheet['D2'].alignment = Alignment(horizontal='center')




        for i in range(14):
            celda = chr(65 + i) + '4'
            sheet[celda] = self.data_completa[i]


        #path_desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        #print(path_desktop)

        wb.save('Indicadores_financieros.xlsx')
        #wb.save(path_desktop)
        print(self.data_completa)

